# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Convert the markdown tables of a Hugo content page into an Excel workbook.

Usage: uv run tools/md-tables-to-xlsx.py <content page.md> <output.xlsx>

Every heading (up to level 3) that has tables below it becomes a sheet.
Each table becomes a named Excel table object; notes and deeper headings
between the tables are written as plain text rows in document order.
Other page prose is left out.
"""

import re
import sys
import tomllib
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import ExcelWriter

REPO_ROOT = Path(__file__).resolve().parent.parent
SITE_CONFIG = REPO_ROOT / "config/_default/hugo.toml"
CONTENT_ROOT = REPO_ROOT / "content"
FRONT_MATTER_DELIMITER = "+++"
SHEET_HEADING_LEVEL = 3
FIXED_TIMESTAMP = datetime(2000, 1, 1)
ZIP_DATE_TIME = (1980, 1, 1, 0, 0, 0)
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 40
NOTE_COLUMN_WIDTH = 40
NOTE_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TABLE_STYLE = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
INLINE_MARKUP = [
    (re.compile(r"\{\{<\s*abbr\s+(\S+)\s*>\}\}"), r"\1"),
    (re.compile(r"\[\^[^\]]+\]"), ""),
    (re.compile(r"\[([^\]]+)\]\([^)]*\)"), r"\1"),
    (re.compile(r"`([^`]*)`"), r"\1"),
    (re.compile(r"\*\*(.+?)\*\*"), r"\1"),
    (re.compile(r"\*(.+?)\*"), r"\1"),
]
CELL_MARKUP = INLINE_MARKUP[:4]
ALIGNMENTS = {(True, True): "center", (True, False): "left", (False, True): "right", (False, False): None}


@dataclass
class Heading:
    level: int
    text: str


@dataclass
class Note:
    text: str


@dataclass
class MdTable:
    header: list[str]
    alignments: list[str | None]
    rows: list[list[str]]


Block = Heading | Note | MdTable


@dataclass
class Section:
    title: str
    blocks: list[Block]


@dataclass
class WrittenTable:
    sheet: str
    name: str
    rows: int


def plain(text: str, patterns=INLINE_MARKUP) -> str:
    for pattern, replacement in patterns:
        text = pattern.sub(replacement, text)
    return text.strip()


def split_front_matter(text: str) -> tuple[dict, str]:
    lines = text.splitlines()
    if not lines or lines[0].strip() != FRONT_MATTER_DELIMITER:
        raise ValueError("page must start with TOML front matter")
    end = lines.index(FRONT_MATTER_DELIMITER, 1)
    return tomllib.loads("\n".join(lines[1:end])), "\n".join(lines[end + 1:])


def split_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|") and not line.endswith("\\|"):
        line = line[:-1]
    return [cell.strip().replace("\\|", "|") for cell in re.split(r"(?<!\\)\|", line)]


def is_alignment_row(line: str) -> bool:
    cells = split_row(line)
    return line.strip().startswith("|") and all(re.fullmatch(r":?-+:?", cell) for cell in cells)


def alignment_of(cell: str) -> str | None:
    return ALIGNMENTS[(cell.startswith(":"), cell.endswith(":"))]


def is_block_start(line: str) -> bool:
    stripped = line.strip()
    return not stripped or stripped.startswith(("#", ">", "|", "{{<", "[^"))


def parse_blocks(body: str) -> list[Block]:
    blocks: list[Block] = []
    lines = body.splitlines()
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if not stripped or stripped.startswith(("{{<", "[^")):
            i += 1
        elif match := re.fullmatch(r"(#{1,6})\s+(.*)", stripped):
            blocks.append(Heading(len(match.group(1)), plain(match.group(2))))
            i += 1
        elif stripped.startswith(">"):
            note_lines = []
            while i < len(lines) and lines[i].strip().startswith(">"):
                content = lines[i].strip().lstrip(">").strip()
                if not re.fullmatch(r"\[![A-Z]+\]", content):
                    note_lines.append(content)
                i += 1
            blocks.append(Note(plain(" ".join(note_lines))))
        elif stripped.startswith("|") and i + 1 < len(lines) and is_alignment_row(lines[i + 1]):
            header = [plain(cell, CELL_MARKUP) for cell in split_row(lines[i])]
            alignments = [alignment_of(cell) for cell in split_row(lines[i + 1])]
            i += 2
            rows = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                cells = [plain(cell, CELL_MARKUP) for cell in split_row(lines[i])]
                rows.append((cells + [""] * len(header))[: len(header)])
                i += 1
            blocks.append(MdTable(header, alignments, rows))
        else:
            i += 1
            while i < len(lines) and not is_block_start(lines[i]):
                i += 1
    return blocks


def sections_with_tables(blocks: list[Block]) -> list[Section]:
    sections: list[Section] = []
    current: Section | None = None
    for block in blocks:
        if isinstance(block, Heading) and block.level <= SHEET_HEADING_LEVEL:
            current = Section(block.text, [])
            sections.append(current)
        elif current is not None:
            current.blocks.append(block)
    return [section for section in sections if any(isinstance(b, MdTable) for b in section.blocks)]


def page_language(page: Path) -> str:
    return page.resolve().relative_to(CONTENT_ROOT).parts[0]


def unique(name: str, used: set[str]) -> str:
    candidate = name
    counter = 2
    while candidate in used:
        candidate = f"{name}{counter}"
        counter += 1
    used.add(candidate)
    return candidate


def sheet_title(title: str, used: set[str]) -> str:
    return unique(re.sub(r"[\[\]:*?/\\]", "", title)[:31], used)


def table_name(title: str, used: set[str]) -> str:
    words = [word[:1].upper() + word[1:] for word in title.split()]
    name = re.sub(r"\W", "", "".join(words))
    if not name or name[0].isdigit():
        name = "Table" + name
    return unique(name, used)


def is_rtl(text: str) -> bool:
    return any(unicodedata.bidirectional(char) in ("R", "AL") or 0x0590 <= ord(char) <= 0x08FF for char in text)


def set_text(cell, text: str):
    cell.value = text or None
    if text:
        cell.data_type = "s"
        cell.number_format = "@"
    return cell


def write_table(ws: Worksheet, table: MdTable, start_row: int, name: str, widths: dict[int, int]) -> int:
    headers = []
    used_headers: set[str] = set()
    for column, text in enumerate(table.header, 1):
        header = unique(text or f"Column {column}", used_headers)
        headers.append(header)
        set_text(ws.cell(start_row, column), header)
        widths[column] = max(widths.get(column, 0), len(header))
    for row, cells in enumerate(table.rows, start_row + 1):
        for column, text in enumerate(cells, 1):
            cell = set_text(ws.cell(row, column), text)
            cell.alignment = Alignment(horizontal=table.alignments[column - 1], readingOrder=2 if is_rtl(text) else 0)
            widths[column] = max(widths.get(column, 0), len(text))
    end_row = start_row + len(table.rows)
    ref = f"A{start_row}:{get_column_letter(len(headers))}{end_row}"
    ws.add_table(Table(displayName=name, ref=ref, tableStyleInfo=TABLE_STYLE))
    return end_row


def write_section(ws: Worksheet, section: Section, used_names: set[str]) -> list[WrittenTable]:
    written = []
    widths: dict[int, int] = {}
    row = 1
    title = section.title
    has_notes = False
    for block in section.blocks:
        if isinstance(block, Heading):
            ws.cell(row, 1, block.text).font = Font(bold=True)
            title = block.text
            row += 2
        elif isinstance(block, Note):
            set_text(ws.cell(row, 1), block.text).alignment = NOTE_ALIGNMENT
            has_notes = True
            row += 2
        else:
            name = table_name(title, used_names)
            row = write_table(ws, block, row, name, widths) + 2
            written.append(WrittenTable(ws.title, name, len(block.rows)))
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
    if has_notes:
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, NOTE_COLUMN_WIDTH)
    return written


def build_workbook(page: Path, site: dict) -> tuple[Workbook, list[WrittenTable]]:
    front_matter, body = split_front_matter(page.read_text(encoding="utf-8"))
    sections = sections_with_tables(parse_blocks(body))
    if not sections:
        raise ValueError(f"{page} contains no markdown tables")

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = front_matter["title"]
    wb.properties.description = front_matter.get("description", "")
    wb.properties.creator = site["title"]
    wb.properties.lastModifiedBy = site["title"]
    wb.properties.language = page_language(page)
    wb.properties.created = FIXED_TIMESTAMP
    wb.properties.modified = FIXED_TIMESTAMP

    used_sheets: set[str] = set()
    used_tables: set[str] = set()
    written: list[WrittenTable] = []
    for section in sections:
        ws = wb.create_sheet(sheet_title(section.title, used_sheets))
        written.extend(write_section(ws, section, used_tables))
    return wb, written


def save_reproducibly(wb: Workbook, path: Path) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        ExcelWriter(wb, archive).save()
    with zipfile.ZipFile(path) as source:
        entries = [(info.filename, source.read(info)) for info in source.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=ZIP_DATE_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            target.writestr(info, data)


def rows_in(ref: str) -> int:
    start, end = ref.split(":")
    return int(re.sub(r"[A-Z]", "", end)) - int(re.sub(r"[A-Z]", "", start))


def verify(path: Path, expected: list[WrittenTable]) -> None:
    wb = load_workbook(path)
    found = [WrittenTable(ws.title, name, rows_in(ws.tables[name].ref)) for ws in wb.worksheets for name in ws.tables]
    if found != expected:
        raise AssertionError(f"workbook mismatch:\n  expected {expected}\n  found    {found}")


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print(__doc__.strip(), file=sys.stderr)
        return 2
    page, output = Path(argv[1]), Path(argv[2])
    site = tomllib.loads(SITE_CONFIG.read_text(encoding="utf-8"))
    wb, written = build_workbook(page, site)
    output.parent.mkdir(parents=True, exist_ok=True)
    save_reproducibly(wb, output)
    verify(output, written)
    for table in written:
        print(f"{table.sheet}: {table.name} ({table.rows} rows)")
    print(f"wrote {output}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
